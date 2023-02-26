import openpyxl
import openai
from config import API_KEY
import csv
from collections import OrderedDict


openai.api_key = API_KEY
filename = 'reviews'


def read_reviews_from_xlsx(filename):
    # Открываем файл xlsx и получаем первый лист
    wb = openpyxl.load_workbook(f'{filename}.xlsx')
    ws = wb.active

    # Получаем значения из первой строки как заголовки
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(row=1, column=col).value)

    # Инициализируем словарь отзывов
    reviews = {}

    # Итерируемся по строкам и добавляем данные в словарь
    for row in range(2, ws.max_row + 1):
        review_dict = {}
        for col in range(1, ws.max_column + 1):
            review_dict[headers[col - 1]] = ws.cell(row=row, column=col).value
        reviews[f'{row-1}'] = review_dict

    # Возвращаем словарь отзывов
    return reviews


def get_reviews_list(reviews):
    reviews_list = []
    for k, v in reviews.items():
        reviews_list.append(v['review text'])
    return reviews_list


def write_csv(filename, data):
    with open(f'{filename}__analyzed.csv', 'w', newline='') as file:
        fieldnames = ['email', 'review text', 'date', 'rate']
        writer = csv.DictWriter(file, fieldnames=fieldnames)

        writer.writeheader()  # записываем заголовок

        for item in data:
            writer.writerow(item)


def sort_reviews_by_sentiment(reviews):
    """
    Принимает список отзывов и сортирует его по положительности контекста с помощью API OpenAI.
    Возвращает отсортированный список отзывов.
    """
    # Создаем список для хранения результатов
    results = []

    # Для каждого отзыва в списке
    for review in reviews:
        # Отправляем запрос в OpenAI API, чтобы определить сентимент отзыва
        response = openai.Completion.create(
            engine="text-davinci-003",
            prompt=f"Sort this review based on its sentiment:\n{review}\nPositive:\nNegative:",
            temperature=0.5,
            max_tokens=1,
            n=1,
            stop=None,
            presence_penalty=0.5,
        )
        # Получаем ответ
        sentiment = response.choices[0].text
        # Добавляем отзыв и его сентимент в список результатов
        results.append((review, sentiment))

    # Сортируем список результатов по сентименту в порядке убывания
    sorted_results = sorted(results, key=lambda x: x[1], reverse=True)

    # Создаем список только из отзывов, без сентимента
    sorted_reviews = [x[0] for x in sorted_results]
    sorted_reviews = list(OrderedDict.fromkeys(sorted_reviews))
    # Возвращаем отсортированный список отзывов
    return sorted_reviews


# def sort_reviews_by_sentiment(reviews):
#     """
#     Принимает список отзывов и сортирует его по степени положительности контекста с помощью API ChatGPT.
#     Возвращает отсортированный список отзывов.
#     """
#     # Создаем список для хранения результатов
#     results = [(review, 0) for review in reviews]
#
#     # Для каждой пары отзывов
#     for i in range(len(reviews)):
#         for j in range(i + 1, len(reviews)):
#             # Отправляем запрос в ChatGPT API, чтобы сравнить пару отзывов
#             prompt = f"Which review is more positive?\n1. {reviews[i]}\n2. {reviews[j]}\n"
#             response = openai.Completion.create(
#                 engine="text-davinci-003",
#                 prompt=prompt,
#                 temperature=0.5,
#                 max_tokens=1,
#                 n=1,
#                 stop=None,
#             )
#             # Получаем ответ в формате JSON
#             comparison = response.choices[0].text
#             # Если первый отзыв более положительный
#             if comparison == "1":
#                 # Увеличиваем счетчик положительности первого отзыва
#                 results[i] = (reviews[i], results[i][1] + 1)
#             # Если второй отзыв более положительный
#             elif comparison == "2":
#                 # Увеличиваем счетчик положительности второго отзыва
#                 results[j] = (reviews[j], results[j][1] + 1)
#             # Если отзывы равноценны
#             else:
#                 # Ничего не делаем
#                 pass
#
#     # Сортируем список результатов по счетчику положительности в порядке убывания
#     sorted_results = sorted(results, key=lambda x: x[1], reverse=True)
#
#     # Создаем список только из отзывов, без счетчиков положительности
#     sorted_reviews = [x[0] for x in sorted_results]
#
#     # Возвращаем отсортированный список отзывов
#     print(sorted_reviews)
#     return sorted_reviews


def main():
    reviews = read_reviews_from_xlsx(filename)
    reviews_list = get_reviews_list(reviews)
    sorted_reviews_list = list(enumerate(sort_reviews_by_sentiment(reviews_list), 1))
    sorted_reviews_list = reversed(sorted_reviews_list)
    sorted_reviews = []
    for rev in sorted_reviews_list:
        for k, v in reviews.items():
            if rev[1] in v.values():
                print(v)
                v['rate'] = rev[0]
                sorted_reviews.append(v)
                # print(sorted_reviews)
    write_csv(filename, sorted_reviews)


if __name__ == '__main__':
    main()


